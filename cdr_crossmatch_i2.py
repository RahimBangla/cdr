#!/usr/bin/env python3
"""CDR occurrence investigation: cross-match + i2-style relation graph PDF."""

from __future__ import annotations

import math
import re
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import networkx as nx
import openpyxl
from matplotlib.patches import Circle, FancyBboxPatch, FancyArrowPatch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    Image,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

warnings.filterwarnings("ignore")

BASE = Path(r"C:\Users\rahim\Downloads\cdr")
OUT_DIR = BASE / "output"
OUT_DIR.mkdir(exist_ok=True)

# Service / shortcodes / non-subscriber parties to de-emphasize in graphs
NOISE_PATTERNS = (
    re.compile(r"^[A-Za-z]"),  # alpha names e.g. GP Point
    re.compile(r"^446"),  # hex/hex-encoded short SMS bodies often appear as BPARTY
    re.compile(r"^4750"),  # GP short/service patterns
    re.compile(r"^\d{1,6}$"),  # very short codes
)


def norm_msisdn(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return None
    # Keep alphanumeric short labels as-is (service names)
    if re.search(r"[A-Za-z]", s):
        return s
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    if digits.startswith("880") and len(digits) >= 13:
        return digits
    if digits.startswith("0") and len(digits) == 11:
        return "880" + digits[1:]
    if len(digits) == 10 and digits[0] == "1":
        return "880" + digits
    return digits


def is_subscriber(msisdn: str | None) -> bool:
    if not msisdn:
        return False
    if any(p.search(msisdn) for p in NOISE_PATTERNS):
        return False
    # Pure MSISDN only — reject hex/SMS payload fragments (e.g. ...656C)
    if not re.fullmatch(r"8801\d{9}", msisdn):
        return False
    return True


def parse_dt(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ("%Y%m%d%H%M%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:14] if fmt == "%Y%m%d%H%M%S" else s, fmt)
        except ValueError:
            continue
    return None


def classify_usage(usage: str | None) -> str:
    u = (usage or "").upper().replace(" ", "").replace("-", "").replace("_", "")
    if "SMS" in u:
        direction = "MT" if "MT" in u else ("MO" if "MO" in u else "")
        return f"SMS{direction}" if direction else "SMS"
    if any(x in u for x in ("CALL", "MOC", "MTC", "VOICE")):
        if "MTC" in u or u.endswith("MT") or "CALLMT" in u:
            return "CALL_MT"
        if "MOC" in u or "CALLMO" in u or u.endswith("MO"):
            return "CALL_MO"
        return "CALL"
    if "GPRS" in u or "DATA" in u or "INTERNET" in u:
        return "DATA"
    return usage or "OTHER"


@dataclass
class Record:
    source_file: str
    sheet: str
    target: str
    start: datetime | None
    provider: str
    aparty: str
    bparty: str
    duration: int
    usage: str
    usage_class: str
    network: str
    lac: str
    ci: str
    imei: str
    imsi: str
    address: str


@dataclass
class Investigation:
    records: list[Record] = field(default_factory=list)
    targets: set[str] = field(default_factory=set)


HEADER_MAP = {
    "START_DTTIME": "start",
    "START": "start",
    "PROVIDER_NAME": "provider",
    "PROVIDER NAME": "provider",
    "APARTY": "aparty",
    "BPARTY": "bparty",
    "CALL_DURATION": "duration",
    "CALL DURATION": "duration",
    "USAGE_TYPE": "usage",
    "USAGE TYPE": "usage",
    "NETWORK_TYPE": "network",
    "NETWORK TYPE": "network",
    "LACSTARTA": "lac",
    "CISTARTA": "ci",
    "IMEI": "imei",
    "IMSI": "imsi",
    "IMSIA": "imsi",
    "ADDRESS": "address",
}


def load_all(base: Path) -> Investigation:
    inv = Investigation()
    for path in sorted(base.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() if h is not None else f"COL{i}" for i, h in enumerate(raw_headers)]
            col = {}
            for i, h in enumerate(headers):
                key = HEADER_MAP.get(h.upper() if isinstance(h, str) else h)
                if key:
                    col[key] = i + 1
            if "aparty" not in col or "bparty" not in col:
                continue

            # Infer target from sheet name or first aparty
            sheet_target = None
            m = re.search(r"(880\d{10}|\d{10,13})", sheet_name)
            if m:
                sheet_target = norm_msisdn(m.group(1))
            if not sheet_target:
                m = re.search(r"(880\d{10}|\d{10,13})", path.stem)
                if m:
                    sheet_target = norm_msisdn(m.group(1))

            for r in range(2, ws.max_row + 1):
                aparty = norm_msisdn(ws.cell(r, col["aparty"]).value)
                bparty = norm_msisdn(ws.cell(r, col["bparty"]).value)
                if not aparty:
                    continue
                target = sheet_target or aparty
                inv.targets.add(target)
                dur_raw = ws.cell(r, col["duration"]).value if "duration" in col else 0
                try:
                    duration = int(float(dur_raw or 0))
                except (TypeError, ValueError):
                    duration = 0
                usage = str(ws.cell(r, col["usage"]).value or "") if "usage" in col else ""
                inv.records.append(
                    Record(
                        source_file=path.name,
                        sheet=sheet_name,
                        target=target,
                        start=parse_dt(ws.cell(r, col["start"]).value) if "start" in col else None,
                        provider=str(ws.cell(r, col["provider"]).value or "") if "provider" in col else "",
                        aparty=aparty,
                        bparty=bparty or "",
                        duration=duration,
                        usage=usage,
                        usage_class=classify_usage(usage),
                        network=str(ws.cell(r, col["network"]).value or "") if "network" in col else "",
                        lac=str(ws.cell(r, col["lac"]).value or "") if "lac" in col else "",
                        ci=str(ws.cell(r, col["ci"]).value or "") if "ci" in col else "",
                        imei=str(ws.cell(r, col["imei"]).value or "").split(".")[0] if "imei" in col else "",
                        imsi=str(ws.cell(r, col["imsi"]).value or "").split(".")[0] if "imsi" in col else "",
                        address=str(ws.cell(r, col["address"]).value or "").strip() if "address" in col else "",
                    )
                )
        wb.close()
    return inv


def fmt_msisdn(n: str) -> str:
    digits = re.sub(r"\D", "", n)
    if digits.startswith("880") and len(digits) == 13:
        return f"+{digits[:3]} {digits[3:5]} {digits[5:9]} {digits[9:]}"
    return n


def short(n: str, width: int = 14) -> str:
    if len(n) <= width:
        return n
    return n[: width - 1] + "…"


def fmt_duration(seconds: int | float | None) -> str:
    s = int(seconds or 0)
    if s <= 0:
        return "0s"
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    if h:
        return f"{h}h{m:02d}m{sec:02d}s"
    if m:
        return f"{m}m{sec:02d}s"
    return f"{sec}s"


def edge_label(events: int, duration: int) -> str:
    if duration > 0:
        return f"{events} | {fmt_duration(duration)}"
    return str(events)


def analyze(inv: Investigation) -> dict:
    targets = sorted(inv.targets)
    by_target: dict[str, list[Record]] = defaultdict(list)
    for rec in inv.records:
        by_target[rec.target].append(rec)

    # Contact sets (subscriber B-parties only)
    contacts: dict[str, Counter] = {t: Counter() for t in targets}
    call_links: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"calls": 0, "sms": 0, "duration": 0, "first": None, "last": None, "types": Counter()}
    )
    imei_by_target: dict[str, Counter] = {t: Counter() for t in targets}
    imsi_by_target: dict[str, Counter] = {t: Counter() for t in targets}
    cell_by_target: dict[str, Counter] = {t: Counter() for t in targets}
    addr_by_target: dict[str, Counter] = {t: Counter() for t in targets}

    for t, rows in by_target.items():
        for rec in rows:
            b = rec.bparty
            if is_subscriber(b):
                contacts[t][b] += 1
                a, bsorted = sorted([t, b])
                link = call_links[(a, bsorted)]
                if "SMS" in rec.usage_class:
                    link["sms"] += 1
                else:
                    link["calls"] += 1
                # Always accumulate CDR duration (voice seconds; SMS usually 0)
                link["duration"] += max(0, int(rec.duration or 0))
                link["types"][rec.usage_class] += 1
                if rec.start:
                    if link["first"] is None or rec.start < link["first"]:
                        link["first"] = rec.start
                    if link["last"] is None or rec.start > link["last"]:
                        link["last"] = rec.start
            if rec.imei and rec.imei not in {"0", "None"}:
                imei_by_target[t][rec.imei] += 1
            if rec.imsi and rec.imsi not in {"0", "None"}:
                imsi_by_target[t][rec.imsi] += 1
            if rec.lac and rec.ci:
                cell_by_target[t][f"{rec.lac}/{rec.ci}"] += 1
            if rec.address:
                addr_by_target[t][rec.address] += 1

    # Target-to-target matrix
    t2t = []
    for i, a in enumerate(targets):
        for b in targets[i + 1 :]:
            common = set(contacts[a]) & set(contacts[b])
            common.discard(a)
            common.discard(b)
            key = tuple(sorted([a, b]))
            link = call_links.get(key, {"calls": 0, "sms": 0, "duration": 0})
            # also count from each target's CDR when BPARTY is the other target
            dir_a = contacts[a][b]
            dir_b = contacts[b][a]
            t2t.append(
                {
                    "a": a,
                    "b": b,
                    "direct_ab": dir_a,
                    "direct_ba": dir_b,
                    "direct_total": dir_a + dir_b,
                    "common_contacts": sorted(common, key=lambda x: -(contacts[a][x] + contacts[b][x])),
                    "common_count": len(common),
                    "duration": int(link.get("duration", 0) or 0),
                    "sms": int(link.get("sms", 0) or 0),
                    "voice_calls": int(link.get("calls", 0) or 0),
                }
            )

    # Common contacts across N+ targets
    contact_owners: dict[str, set[str]] = defaultdict(set)
    contact_hits: dict[str, int] = Counter()
    for t, ctr in contacts.items():
        for num, cnt in ctr.items():
            if num in inv.targets:
                continue
            contact_owners[num].add(t)
            contact_hits[num] += cnt
    multi_common = []
    for num, owners in contact_owners.items():
        if len(owners) < 2:
            continue
        per_dur = {}
        total_dur = 0
        for t in sorted(owners):
            d = int(call_links.get(tuple(sorted([t, num])), {}).get("duration", 0) or 0)
            per_dur[t] = d
            total_dur += d
        multi_common.append(
            {
                "number": num,
                "targets": sorted(owners),
                "target_count": len(owners),
                "total_hits": contact_hits[num],
                "total_duration": total_dur,
                "per_target": {t: contacts[t][num] for t in sorted(owners)},
                "per_target_duration": per_dur,
            }
        )
    multi_common.sort(key=lambda x: (-x["target_count"], -x["total_hits"], -x["total_duration"]))

    # Shared IMEI across targets (same handset / SIM swap)
    imei_owners: dict[str, set[str]] = defaultdict(set)
    for t, ctr in imei_by_target.items():
        for imei in ctr:
            if imei:
                imei_owners[imei].add(t)
    shared_imei = [
        {"imei": imei, "targets": sorted(owners), "counts": {t: imei_by_target[t][imei] for t in owners}}
        for imei, owners in imei_owners.items()
        if len(owners) >= 2
    ]
    shared_imei.sort(key=lambda x: -len(x["targets"]))

    # Shared cells
    cell_owners: dict[str, set[str]] = defaultdict(set)
    for t, ctr in cell_by_target.items():
        for cell in ctr:
            cell_owners[cell].add(t)
    shared_cells = [
        {
            "cell": cell,
            "targets": sorted(owners),
            "counts": {t: cell_by_target[t][cell] for t in owners},
            "address_hint": "",
        }
        for cell, owners in cell_owners.items()
        if len(owners) >= 2
    ]
    # attach sample address
    for item in shared_cells:
        lac, _, ci = item["cell"].partition("/")
        for rec in inv.records:
            if rec.lac == lac and rec.ci == ci and rec.address:
                item["address_hint"] = rec.address
                break
    shared_cells.sort(key=lambda x: (-len(x["targets"]), -sum(x["counts"].values())))

    # Target summaries
    summaries = []
    for t in targets:
        rows = by_target[t]
        dates = [r.start for r in rows if r.start]
        top_contacts = contacts[t].most_common(15)
        imeis = imei_by_target[t].most_common(3)
        summaries.append(
            {
                "target": t,
                "records": len(rows),
                "provider": Counter(r.provider for r in rows if r.provider).most_common(1),
                "unique_contacts": len(contacts[t]),
                "period": (min(dates), max(dates)) if dates else (None, None),
                "top_contacts": top_contacts,
                "imeis": imeis,
                "imsis": imsi_by_target[t].most_common(2),
                "top_cells": cell_by_target[t].most_common(5),
                "top_addrs": addr_by_target[t].most_common(3),
                "usage": Counter(r.usage_class for r in rows),
            }
        )

    # Graph edges for i2: targets + strong common contacts + direct target links
    # tuple: (a, b, weight/events, label, kind, duration_seconds)
    graph_nodes = set(targets)
    graph_edges = []

    for item in t2t:
        if item["direct_total"] > 0:
            dur = int(item.get("duration", 0) or 0)
            graph_edges.append(
                (
                    item["a"],
                    item["b"],
                    item["direct_total"],
                    edge_label(item["direct_total"], dur),
                    "target_link",
                    dur,
                )
            )

    # include common contacts linked to 2+ targets (top by hits)
    for item in multi_common[:40]:
        graph_nodes.add(item["number"])
        for t in item["targets"]:
            w = item["per_target"][t]
            dur = int(item["per_target_duration"].get(t, 0) or 0)
            graph_edges.append((t, item["number"], w, edge_label(w, dur), "common", dur))

    # plus top exclusive contacts per target (to show neighbourhood)
    for t in targets:
        added = 0
        for num, cnt in contacts[t].most_common(25):
            if num in graph_nodes and num not in targets:
                continue
            if len(contact_owners.get(num, set())) >= 2:
                continue  # already as common
            if cnt < 3:
                continue
            graph_nodes.add(num)
            dur = int(call_links.get(tuple(sorted([t, num])), {}).get("duration", 0) or 0)
            graph_edges.append((t, num, cnt, edge_label(cnt, dur), "exclusive", dur))
            added += 1
            if added >= 6:
                break

    # IMEI nodes for shared handsets
    imei_nodes = []
    for item in shared_imei[:10]:
        node = f"IMEI:{item['imei'][-8:]}"
        imei_nodes.append((node, item))
        graph_nodes.add(node)
        for t in item["targets"]:
            graph_edges.append((t, node, item["counts"][t], "handset", "imei", 0))

    return {
        "targets": targets,
        "by_target": by_target,
        "contacts": contacts,
        "t2t": t2t,
        "multi_common": multi_common,
        "shared_imei": shared_imei,
        "shared_cells": shared_cells[:50],
        "summaries": summaries,
        "graph_nodes": graph_nodes,
        "graph_edges": graph_edges,
        "imei_nodes": imei_nodes,
        "call_links": call_links,
    }


# ---------------- i2-style charts ----------------

TARGET_COLOR = "#C0392B"
COMMON_COLOR = "#2980B9"
EXCLUSIVE_COLOR = "#27AE60"
IMEI_COLOR = "#8E44AD"
EDGE_TARGET = "#922B21"
EDGE_COMMON = "#2471A3"
EDGE_EXCL = "#1E8449"
EDGE_IMEI = "#6C3483"


def draw_i2_graph(analysis: dict, out_path: Path) -> Path:
    G = nx.Graph()
    targets = set(analysis["targets"])
    for n in analysis["graph_nodes"]:
        if n.startswith("IMEI:"):
            kind = "imei"
        elif n in targets:
            kind = "target"
        else:
            owners = sum(1 for e in analysis["graph_edges"] if n in e[:2] and e[4] == "common")
            kind = "common" if owners else "exclusive"
        G.add_node(n, kind=kind)

    for edge in analysis["graph_edges"]:
        a, b, w, label, kind = edge[0], edge[1], edge[2], edge[3], edge[4]
        dur = int(edge[5]) if len(edge) > 5 else 0
        if G.has_edge(a, b):
            G[a][b]["weight"] += w
            G[a][b]["duration"] = G[a][b].get("duration", 0) + dur
            G[a][b]["label"] = edge_label(int(G[a][b]["weight"]), int(G[a][b]["duration"]))
        else:
            G.add_edge(a, b, weight=w, label=label, kind=kind, duration=dur)

    if len(G) == 0:
        fig, ax = plt.subplots(figsize=(14, 10))
        ax.text(0.5, 0.5, "No graphable links", ha="center")
        fig.savefig(out_path, dpi=160, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        return out_path

    # Clean 3-ring shell layout (no spring scramble — avoids line pile-ups)
    by_kind = {"target": [], "common": [], "exclusive": [], "imei": []}
    for n in G.nodes:
        by_kind[G.nodes[n]["kind"]].append(n)
    for k in by_kind:
        by_kind[k].sort()

    pos = {}
    shells = [
        (by_kind["target"], 1.45, 0.0),
        (by_kind["common"], 3.05, 0.28),
        (by_kind["exclusive"] + by_kind["imei"], 4.55, 0.55),
    ]
    for members, radius, phase in shells:
        count = max(len(members), 1)
        for i, n in enumerate(members):
            ang = phase + (2 * math.pi * i / count) - math.pi / 2
            pos[n] = (radius * math.cos(ang), radius * math.sin(ang))

    fig, ax = plt.subplots(figsize=(16.5, 12.5))
    ax.set_facecolor("#F7F4EF")
    fig.patch.set_facecolor("#F7F4EF")

    ax.add_patch(
        FancyBboxPatch(
            (-0.02, -0.02),
            1.04,
            1.04,
            boxstyle="round,pad=0.01,rounding_size=0.01",
            transform=ax.transAxes,
            fill=False,
            edgecolor="#5D6D7E",
            linewidth=1.2,
            zorder=0,
        )
    )

    # Curved edges with alternating bow so lines don't stack on one stroke
    edge_list = sorted(G.edges(data=True), key=lambda x: -x[2].get("weight", 1))
    for i, (u, v, data) in enumerate(edge_list):
        kind = data.get("kind", "exclusive")
        color = {
            "target_link": EDGE_TARGET,
            "common": EDGE_COMMON,
            "imei": EDGE_IMEI,
        }.get(kind, EDGE_EXCL)
        w = data.get("weight", 1)
        dur = int(data.get("duration", 0) or 0)
        lw = 0.9 + min(4.2, math.log1p(w) * 0.9)
        rad = (0.12 + (i % 7) * 0.04) * (1 if i % 2 == 0 else -1)
        ax.add_patch(
            FancyArrowPatch(
                pos[u],
                pos[v],
                connectionstyle=f"arc3,rad={rad:.3f}",
                arrowstyle="-",
                mutation_scale=1,
                linewidth=lw,
                color=color,
                alpha=0.55 if kind != "target_link" else 0.8,
                shrinkA=10,
                shrinkB=10,
                zorder=1,
            )
        )
        # Labels only on key links to avoid text collisions
        if kind == "target_link" or w >= 25 or dur >= 1800:
            x1, y1 = pos[u]
            x2, y2 = pos[v]
            # offset label perpendicular to chord along the bow
            mx, my = (x1 + x2) / 2, (y1 + y2) / 2
            dx, dy = x2 - x1, y2 - y1
            length = math.hypot(dx, dy) or 1
            ox, oy = -dy / length * (0.18 * abs(rad) / 0.12), dx / length * (0.18 * abs(rad) / 0.12)
            ax.text(
                mx + ox,
                my + oy,
                edge_label(int(w), dur),
                fontsize=6.5,
                color="#1C2833",
                ha="center",
                va="center",
                bbox=dict(boxstyle="round,pad=0.12", fc="#F7F4EF", ec="none", alpha=0.92),
                zorder=2,
            )

    # Nodes
    for n, (x, y) in pos.items():
        kind = G.nodes[n]["kind"]
        if kind == "target":
            face, edge, size, shape = TARGET_COLOR, "#7B241C", 0.22, "circle"
            label = fmt_msisdn(n)
            fs = 8
            fontweight = "bold"
        elif kind == "common":
            face, edge, size, shape = COMMON_COLOR, "#1A5276", 0.16, "circle"
            label = fmt_msisdn(n) if is_subscriber(n) else n
            fs = 6.5
            fontweight = "normal"
        elif kind == "imei":
            face, edge, size, shape = IMEI_COLOR, "#5B2C6F", 0.17, "square"
            label = n
            fs = 6.5
            fontweight = "bold"
        else:
            face, edge, size, shape = EXCLUSIVE_COLOR, "#145A32", 0.13, "circle"
            label = fmt_msisdn(n) if is_subscriber(n) else n
            fs = 6
            fontweight = "normal"

        if shape == "square":
            ax.add_patch(
                FancyBboxPatch(
                    (x - size, y - size),
                    size * 2,
                    size * 2,
                    boxstyle="round,pad=0.02,rounding_size=0.04",
                    facecolor=face,
                    edgecolor=edge,
                    linewidth=1.4,
                    zorder=3,
                )
            )
        else:
            ax.add_patch(Circle((x, y), size, facecolor=face, edgecolor=edge, linewidth=1.6, zorder=3))
            # phone glyph hint
            ax.plot([x], [y], marker="s", markersize=4, color="white", zorder=4)

        ax.text(
            x,
            y - size - 0.08,
            label,
            fontsize=fs,
            ha="center",
            va="top",
            color="#1B2631",
            fontweight=fontweight,
            zorder=5,
        )

    # Legend
    legend_items = [
        (TARGET_COLOR, "Subject (A-Party / Target)"),
        (COMMON_COLOR, "Common contact (2+ targets)"),
        (EXCLUSIVE_COLOR, "Strong exclusive contact"),
        (IMEI_COLOR, "Shared IMEI / Handset"),
    ]
    lx, ly = 0.02, 0.98
    ax.text(lx, ly, "i2-style Association Chart", transform=ax.transAxes, fontsize=12, fontweight="bold", va="top")
    ax.text(
        lx,
        ly - 0.045,
        "Clean shell layout  |  Curved links  |  Labels = key links only (events | duration)",
        transform=ax.transAxes,
        fontsize=8,
        color="#34495E",
        va="top",
    )
    for i, (c, text) in enumerate(legend_items):
        yy = ly - 0.095 - i * 0.035
        ax.add_patch(Circle((lx + 0.012, yy), 0.008, transform=ax.transAxes, facecolor=c, edgecolor="none"))
        ax.text(lx + 0.028, yy, text, transform=ax.transAxes, fontsize=8, va="center", color="#212F3D")

    ax.set_aspect("equal")
    ax.axis("off")
    # padding
    xs = [p[0] for p in pos.values()]
    ys = [p[1] for p in pos.values()]
    pad = 0.85
    ax.set_xlim(min(xs) - pad, max(xs) + pad)
    ax.set_ylim(min(ys) - pad, max(ys) + pad)

    fig.savefig(out_path, dpi=170, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    return out_path


def draw_matrix_heatmap(analysis: dict, out_path: Path) -> Path:
    targets = analysis["targets"]
    n = len(targets)
    mat = [[0 for _ in range(n)] for _ in range(n)]
    label = [t[-10:] for t in targets]
    idx = {t: i for i, t in enumerate(targets)}
    for item in analysis["t2t"]:
        i, j = idx[item["a"]], idx[item["b"]]
        mat[i][j] = item["common_count"]
        mat[j][i] = item["common_count"]
        # put direct on diagonal-off with negative? better separate: add direct*1000 visually no
    # Direct overlay matrix
    direct = [[0 for _ in range(n)] for _ in range(n)]
    for item in analysis["t2t"]:
        i, j = idx[item["a"]], idx[item["b"]]
        direct[i][j] = item["direct_total"]
        direct[j][i] = item["direct_total"]

    fig, axes = plt.subplots(1, 2, figsize=(12.5, 5.2))
    for ax, data, title, cmap in [
        (axes[0], mat, "Common Contacts Count", "YlOrRd"),
        (axes[1], direct, "Direct Events Between Targets", "PuBu"),
    ]:
        im = ax.imshow(data, cmap=cmap)
        ax.set_xticks(range(n))
        ax.set_yticks(range(n))
        ax.set_xticklabels(label, rotation=45, ha="right", fontsize=8)
        ax.set_yticklabels(label, fontsize=8)
        ax.set_title(title, fontsize=10, fontweight="bold")
        for i in range(n):
            for j in range(n):
                ax.text(j, i, str(data[i][j]), ha="center", va="center", fontsize=8, color="#1B2631")
        fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    fig.suptitle("Target Cross-Match Matrices", fontsize=12, fontweight="bold")
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return out_path


# ---------------- PDF ----------------


def build_pdf(inv: Investigation, analysis: dict, graph_path: Path, matrix_path: Path, out_pdf: Path) -> Path:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "T",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1B2631"),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    h2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=colors.HexColor("#922B21"),
        spaceBefore=10,
        spaceAfter=6,
    )
    h3 = ParagraphStyle(
        "H3",
        parent=styles["Heading3"],
        fontSize=10,
        textColor=colors.HexColor("#1A5276"),
        spaceBefore=8,
        spaceAfter=4,
    )
    body = ParagraphStyle("B", parent=styles["Normal"], fontSize=8.5, leading=11)
    small = ParagraphStyle("S", parent=styles["Normal"], fontSize=7.5, leading=9.5)
    center = ParagraphStyle("C", parent=body, alignment=TA_CENTER)

    doc = SimpleDocTemplate(
        str(out_pdf),
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
        title="CDR Occurrence Cross-Match & i2 Relation Chart",
    )
    story = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    story.append(Paragraph("OCCURRENCE INVESTIGATION — CDR CROSS-MATCH", title_style))
    story.append(Paragraph(f"i2-style Association Chart & Relation Analysis  |  Generated: {now}", center))
    story.append(Spacer(1, 6))

    targets = analysis["targets"]
    src_files = sorted({r.source_file for r in inv.records})
    overview = [
        ["Metric", "Value"],
        ["Subject / Target numbers", str(len(targets))],
        ["CDR source files", str(len(src_files))],
        ["Total CDR events loaded", f"{len(inv.records):,}"],
        ["Common contacts (shared by ≥2 targets)", str(len(analysis["multi_common"]))],
        ["Shared IMEI / handsets", str(len(analysis["shared_imei"]))],
        ["Shared LAC/CI cells (≥2 targets)", str(len(analysis["shared_cells"]))],
    ]
    t = Table(overview, colWidths=[9 * cm, 16 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2631")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8F9F9")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BFC9CA")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 8))

    story.append(Paragraph("1. Subject Numbers Under Investigation", h2))
    rows = [["#", "MSISDN", "Events", "Unique Contacts", "Provider", "Period", "Top IMEI"]]
    for i, s in enumerate(analysis["summaries"], 1):
        prov = s["provider"][0][0] if s["provider"] else "-"
        p0, p1 = s["period"]
        period = f"{p0:%Y-%m-%d} → {p1:%Y-%m-%d}" if p0 and p1 else "-"
        imei = s["imeis"][0][0] if s["imeis"] else "-"
        rows.append(
            [
                str(i),
                fmt_msisdn(s["target"]),
                f"{s['records']:,}",
                str(s["unique_contacts"]),
                prov,
                period,
                imei,
            ]
        )
    t = Table(rows, colWidths=[1 * cm, 4.2 * cm, 2.2 * cm, 3 * cm, 3.2 * cm, 5.5 * cm, 5.5 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#922B21")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5D8DC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FDEDEC")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(t)

    story.append(Paragraph("2. i2-Style Relation / Association Chart", h2))
    story.append(
        Paragraph(
            "Red = subject targets; Blue = common contacts shared by 2+ subjects; Green = strong exclusive contacts; "
            "Purple = shared IMEI (possible handset/SIM relationship). Edge labels show event counts from CDR.",
            small,
        )
    )
    story.append(Spacer(1, 4))
    story.append(Image(str(graph_path), width=25.5 * cm, height=18.5 * cm))
    story.append(PageBreak())

    story.append(Paragraph("3. Cross-Match Matrices", h2))
    story.append(Image(str(matrix_path), width=24 * cm, height=10 * cm))
    story.append(Spacer(1, 8))

    story.append(Paragraph("4. Target-to-Target Direct Links & Common Contacts", h2))
    rows = [["Target A", "Target B", "A→B", "B→A", "Direct Total", "Call Duration", "Common Contacts", "Top Shared Numbers"]]
    for item in sorted(analysis["t2t"], key=lambda x: (-x["direct_total"], -x["common_count"])):
        top_shared = ", ".join(fmt_msisdn(x) for x in item["common_contacts"][:5]) or "-"
        rows.append(
            [
                Paragraph(fmt_msisdn(item["a"]), small),
                Paragraph(fmt_msisdn(item["b"]), small),
                str(item["direct_ab"]),
                str(item["direct_ba"]),
                str(item["direct_total"]),
                fmt_duration(item.get("duration", 0)),
                str(item["common_count"]),
                Paragraph(top_shared, small),
            ]
        )
    t = Table(rows, colWidths=[3.0 * cm, 3.0 * cm, 1.3 * cm, 1.3 * cm, 1.8 * cm, 2.2 * cm, 2.2 * cm, 9.5 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5D8DC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF5FB")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(t)
    story.append(PageBreak())

    story.append(Paragraph("5. Multi-Target Common Numbers (Priority Leads)", h2))
    story.append(
        Paragraph(
            "Numbers that appear in CDRs of two or more subjects — primary cross-match leads for occurrence linkage.",
            small,
        )
    )
    rows = [["Rank", "Common Number", "Linked Targets", "#Targets", "Hits", "Duration", "Per-Target Counts"]]
    for i, item in enumerate(analysis["multi_common"][:60], 1):
        per = ", ".join(
            f"{fmt_msisdn(t)[-11:]}:{c}/{fmt_duration(item.get('per_target_duration', {}).get(t, 0))}"
            for t, c in item["per_target"].items()
        )
        rows.append(
            [
                str(i),
                Paragraph(fmt_msisdn(item["number"]), small),
                Paragraph("<br/>".join(fmt_msisdn(t) for t in item["targets"]), small),
                str(item["target_count"]),
                str(item["total_hits"]),
                fmt_duration(item.get("total_duration", 0)),
                Paragraph(per, small),
            ]
        )
    if len(rows) == 1:
        rows.append(["-", "None found", "-", "-", "-", "-", "-"])
    t = Table(rows, colWidths=[1.1 * cm, 3.4 * cm, 4.8 * cm, 1.6 * cm, 1.3 * cm, 1.8 * cm, 10.3 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5D8DC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2F8")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(t)
    story.append(PageBreak())

    story.append(Paragraph("6. Shared IMEI / Handset Cross-Match", h2))
    rows = [["IMEI", "Linked Targets", "Event Counts"]]
    for item in analysis["shared_imei"][:30]:
        rows.append(
            [
                item["imei"],
                Paragraph("<br/>".join(fmt_msisdn(t) for t in item["targets"]), small),
                Paragraph(", ".join(f"{fmt_msisdn(t)[-11:]}:{c}" for t, c in item["counts"].items()), small),
            ]
        )
    if len(rows) == 1:
        rows.append(["None — no IMEI shared across subject SIMs", "-", "-"])
    t = Table(rows, colWidths=[5 * cm, 8 * cm, 12 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5B2C6F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5D8DC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EEF8")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(t)

    story.append(Paragraph("7. Shared Cell / Location Footprint (LAC-CI)", h2))
    rows = [["LAC / CI", "#Targets", "Targets", "Sample Address"]]
    for item in analysis["shared_cells"][:40]:
        rows.append(
            [
                item["cell"],
                str(len(item["targets"])),
                Paragraph("<br/>".join(fmt_msisdn(t) for t in item["targets"]), small),
                Paragraph(short(item["address_hint"], 80), small),
            ]
        )
    if len(rows) == 1:
        rows.append(["None", "-", "-", "-"])
    t = Table(rows, colWidths=[3.5 * cm, 1.8 * cm, 7 * cm, 12.5 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E6655")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5D8DC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8F8F5")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(t)
    story.append(PageBreak())

    story.append(Paragraph("8. Subject Profiles — Top Contacts", h2))
    for s in analysis["summaries"]:
        story.append(Paragraph(f"Subject: {fmt_msisdn(s['target'])}", h3))
        usage = ", ".join(f"{k}:{v}" for k, v in s["usage"].most_common())
        story.append(Paragraph(f"Usage mix: {usage}", small))
        rows = [["Rank", "B-Party", "Events", "Note"]]
        for i, (num, cnt) in enumerate(s["top_contacts"][:12], 1):
            note = "SUBJECT" if num in analysis["targets"] else (
                "COMMON" if any(c["number"] == num for c in analysis["multi_common"]) else ""
            )
            rows.append([str(i), fmt_msisdn(num), str(cnt), note])
        t = Table(rows, colWidths=[1.5 * cm, 5 * cm, 2 * cm, 3 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5D8DC")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F7")]),
                ]
            )
        )
        story.append(t)
        story.append(Spacer(1, 6))

    story.append(Paragraph("9. Source Files", h2))
    for f in src_files:
        story.append(Paragraph(f"• {f}", small))
    story.append(Spacer(1, 10))
    story.append(
        Paragraph(
            "Disclaimer: Analytical output from telecom CDR for occurrence investigation. "
            "Service short-codes and non-MSISDN B-parties excluded from association chart where possible. "
            "Verify critical links against original operator CDR before evidentiary use.",
            small,
        )
    )

    doc.build(story)
    return out_pdf


def export_excel(analysis: dict, out_xlsx: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Common_Contacts"
    ws.append(["Common_Number", "Target_Count", "Total_Hits", "Total_Duration_Sec", "Targets", "Per_Target"])
    for item in analysis["multi_common"]:
        ws.append(
            [
                item["number"],
                item["target_count"],
                item["total_hits"],
                item.get("total_duration", 0),
                ", ".join(item["targets"]),
                "; ".join(
                    f"{t}:{c}({item.get('per_target_duration', {}).get(t, 0)}s)"
                    for t, c in item["per_target"].items()
                ),
            ]
        )

    ws2 = wb.create_sheet("Target_to_Target")
    ws2.append(["A", "B", "A_to_B", "B_to_A", "Direct_Total", "Call_Duration_Sec", "Common_Count", "Common_Numbers"])
    for item in analysis["t2t"]:
        ws2.append(
            [
                item["a"],
                item["b"],
                item["direct_ab"],
                item["direct_ba"],
                item["direct_total"],
                item.get("duration", 0),
                item["common_count"],
                ", ".join(item["common_contacts"]),
            ]
        )

    ws3 = wb.create_sheet("Shared_IMEI")
    ws3.append(["IMEI", "Targets", "Counts"])
    for item in analysis["shared_imei"]:
        ws3.append([item["imei"], ", ".join(item["targets"]), str(item["counts"])])

    ws4 = wb.create_sheet("Shared_Cells")
    ws4.append(["LAC_CI", "Targets", "Counts", "Address"])
    for item in analysis["shared_cells"]:
        ws4.append([item["cell"], ", ".join(item["targets"]), str(item["counts"]), item["address_hint"]])

    ws5 = wb.create_sheet("Target_Summary")
    ws5.append(["Target", "Events", "Unique_Contacts", "Top_IMEI", "Period_Start", "Period_End"])
    for s in analysis["summaries"]:
        p0, p1 = s["period"]
        ws5.append(
            [
                s["target"],
                s["records"],
                s["unique_contacts"],
                s["imeis"][0][0] if s["imeis"] else "",
                p0.isoformat(sep=" ") if p0 else "",
                p1.isoformat(sep=" ") if p1 else "",
            ]
        )

    wb.save(out_xlsx)
    return out_xlsx


def main():
    print("Loading CDRs...")
    inv = load_all(BASE)
    print(f"  targets={sorted(inv.targets)}")
    print(f"  records={len(inv.records)}")
    print("Analyzing cross-match...")
    analysis = analyze(inv)
    print(f"  common contacts={len(analysis['multi_common'])}")
    print(f"  shared IMEI={len(analysis['shared_imei'])}")
    print(f"  shared cells={len(analysis['shared_cells'])}")

    graph_path = OUT_DIR / "i2_relation_chart.png"
    matrix_path = OUT_DIR / "crossmatch_matrix.png"
    print("Drawing i2 chart...")
    draw_i2_graph(analysis, graph_path)
    print("Drawing matrices...")
    draw_matrix_heatmap(analysis, matrix_path)

    pdf_path = OUT_DIR / "CDR_Occurrence_CrossMatch_i2_Report.pdf"
    xlsx_path = OUT_DIR / "CDR_CrossMatch_Tables.xlsx"
    print("Building PDF...")
    build_pdf(inv, analysis, graph_path, matrix_path, pdf_path)
    print("Exporting Excel tables...")
    export_excel(analysis, xlsx_path)

    # Console highlight
    print("\n=== KEY FINDINGS ===")
    print("Targets:", ", ".join(analysis["targets"]))
    print("\nDirect target links:")
    for item in sorted(analysis["t2t"], key=lambda x: -x["direct_total"]):
        if item["direct_total"] or item["common_count"]:
            print(
                f"  {item['a']} <-> {item['b']}: direct={item['direct_total']} "
                f"(A→B {item['direct_ab']}, B→A {item['direct_ba']}), "
                f"duration={fmt_duration(item.get('duration', 0))}, common={item['common_count']}"
            )
    print("\nTop common contacts:")
    for item in analysis["multi_common"][:15]:
        print(f"  {item['number']}  targets={item['target_count']} hits={item['total_hits']} -> {item['targets']}")
    if analysis["shared_imei"]:
        print("\nShared IMEI:")
        for item in analysis["shared_imei"][:10]:
            print(f"  {item['imei']} -> {item['targets']}")
    else:
        print("\nShared IMEI: none")

    print(f"\nPDF: {pdf_path}")
    print(f"XLSX: {xlsx_path}")
    print(f"CHART: {graph_path}")


if __name__ == "__main__":
    main()
